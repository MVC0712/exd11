#!C:\python\python3.9\python.exe
# -*- coding: utf-8 -*-
import cgitb
import cgi
import os
import json
import sys
import io
import urllib.parse
import openpyxl
from PIL import Image

width = 23
height = 23

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
cgitb.enable()
data = sys.stdin.read()
params = json.loads(data)

response = {"res": params["n"]}

print('Content-type: text/html\nAccess-Control-Allow-Origin: *\n')
print("\n\n")
print(json.JSONEncoder().encode(response))
print('\n')

if urllib.parse.unquote(params["n"]) =="1":
    wb = openpyxl.load_workbook('EtchingReport1B1.xlsx')
    sheet = wb.get_sheet_by_name('1Bn')
elif urllib.parse.unquote(params["n"]) =="2":
    wb = openpyxl.load_workbook('EtchingReport2B1.xlsx')
    sheet = wb.get_sheet_by_name('2B1')
elif urllib.parse.unquote(params["n"]) =="3":
    wb = openpyxl.load_workbook('EtchingReport3B1.xlsx')
    sheet = wb.get_sheet_by_name('3B1')
elif urllib.parse.unquote(params["n"]) =="4":
    wb = openpyxl.load_workbook('EtchingReport4B1.xlsx')
    sheet = wb.get_sheet_by_name('4B1')

sheet['f3'] = params["press_date_at"]
sheet['o3'] = params["die_number"]
sheet['v4'] = params["actual_billet_quantities"]

img = Image.open("brain.png")
img = img.resize((width,height),Image.NEAREST)
img.save("brain.png")

# picDir = "./brain.png"
# picDir = "../../EtchingPicture/" + params["etcing_file_url"]
img = Image.open("brain.png")
img = openpyxl.drawing.image.PILImage("brain.png")
img.anchor('P59')
sheet.add_image(img)

wb.save("../../Etching/" + str(params["press_date_at"]) + "_" + str(params["actual_billet_quantities"]) + "_" + str(params["die_number"]) + ".xlsx")
